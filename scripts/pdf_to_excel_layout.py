import math
from collections import defaultdict

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Border, Side


PDF_PATH = r"c:\Users\willy\AppData\Roaming\Cursor\User\workspaceStorage\03dd988471397545d0fea24bc03d5ccf\pdfs\49c8803e-b8b0-4a50-b8b4-091c3b5b560b\korekok.pdf"
XLSX_PATH = r"c:\Users\willy\AppData\Roaming\Cursor\User\workspaceStorage\03dd988471397545d0fea24bc03d5ccf\pdfs\49c8803e-b8b0-4a50-b8b4-091c3b5b560b\korekok.xlsx"


def add_coord(coords, value, tol=0.8):
    for c in coords:
        if abs(c - value) <= tol:
            return
    coords.append(value)


def nearest_index(sorted_coords, value):
    best_i = 0
    best_d = float("inf")
    for i, c in enumerate(sorted_coords):
        d = abs(c - value)
        if d < best_d:
            best_d = d
            best_i = i
    return best_i


def bins_crossing(coords, start, end, tol=0.8):
    left = min(start, end)
    right = max(start, end)
    idx = []
    for i in range(len(coords) - 1):
        c0, c1 = coords[i], coords[i + 1]
        if c1 <= left + tol:
            continue
        if c0 >= right - tol:
            continue
        idx.append(i)
    return idx


with pdfplumber.open(PDF_PATH) as pdf:
    page = pdf.pages[0]

    x_coords = []
    y_coords = []

    for ln in page.lines:
        add_coord(x_coords, ln["x0"])
        add_coord(x_coords, ln["x1"])
        add_coord(y_coords, ln["top"])
        add_coord(y_coords, ln["bottom"])

    for rect in page.rects:
        add_coord(x_coords, rect["x0"])
        add_coord(x_coords, rect["x1"])
        add_coord(y_coords, rect["top"])
        add_coord(y_coords, rect["bottom"])

    x_coords = sorted(x_coords)
    y_coords = sorted(y_coords)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "korekok"

    # Scale PDF geometry to practical Excel sizes.
    x_scale = 0.16
    y_scale = 0.34

    for i in range(len(x_coords) - 1):
        width = max(0.1, (x_coords[i + 1] - x_coords[i]) * x_scale)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width

    for i in range(len(y_coords) - 1):
        height = max(2, (y_coords[i + 1] - y_coords[i]) * y_scale)
        ws.row_dimensions[i + 1].height = height

    thin = Side(style="thin", color="000000")

    top_map = defaultdict(set)
    bottom_map = defaultdict(set)
    left_map = defaultdict(set)
    right_map = defaultdict(set)

    # Draw line objects.
    for ln in page.lines:
        is_horizontal = abs(ln["top"] - ln["bottom"]) <= 0.8
        is_vertical = abs(ln["x0"] - ln["x1"]) <= 0.8

        if is_horizontal:
            row_edge = nearest_index(y_coords, ln["top"])
            for ci in bins_crossing(x_coords, ln["x0"], ln["x1"]):
                top_map[(row_edge, ci)].add("t")
                if row_edge > 0:
                    bottom_map[(row_edge - 1, ci)].add("b")
        elif is_vertical:
            col_edge = nearest_index(x_coords, ln["x0"])
            for ri in bins_crossing(y_coords, ln["top"], ln["bottom"]):
                left_map[(ri, col_edge)].add("l")
                if col_edge > 0:
                    right_map[(ri, col_edge - 1)].add("r")

    # Draw rectangle edges.
    for rect in page.rects:
        top_edge = nearest_index(y_coords, rect["top"])
        bottom_edge = nearest_index(y_coords, rect["bottom"])
        left_edge = nearest_index(x_coords, rect["x0"])
        right_edge = nearest_index(x_coords, rect["x1"])

        for ci in bins_crossing(x_coords, rect["x0"], rect["x1"]):
            top_map[(top_edge, ci)].add("t")
            if bottom_edge > 0:
                bottom_map[(bottom_edge - 1, ci)].add("b")

        for ri in bins_crossing(y_coords, rect["top"], rect["bottom"]):
            left_map[(ri, left_edge)].add("l")
            if right_edge > 0:
                right_map[(ri, right_edge - 1)].add("r")

    n_rows = len(y_coords) - 1
    n_cols = len(x_coords) - 1

    for r in range(n_rows):
        for c in range(n_cols):
            top = thin if (r, c) in top_map else None
            bottom = thin if (r, c) in bottom_map else None
            left = thin if (r, c) in left_map else None
            right = thin if (r, c) in right_map else None
            ws.cell(row=r + 1, column=c + 1).border = Border(
                top=top, bottom=bottom, left=left, right=right
            )

    # Place text.
    texts = defaultdict(list)
    words = page.extract_words(
        x_tolerance=1.5,
        y_tolerance=1.5,
        keep_blank_chars=False,
        use_text_flow=True,
    )

    for w in words:
        cx = (w["x0"] + w["x1"]) / 2
        cy = (w["top"] + w["bottom"]) / 2
        ci = min(max(nearest_index(x_coords, cx), 0), n_cols - 1)
        ri = min(max(nearest_index(y_coords, cy), 0), n_rows - 1)
        texts[(ri, ci)].append((w["x0"], w["top"], w["text"]))

    for (ri, ci), items in texts.items():
        items.sort(key=lambda t: (round(t[1], 1), t[0]))
        lines = []
        current_y = None
        current_words = []
        for x, y, txt in items:
            if current_y is None or abs(y - current_y) <= 1.2:
                current_words.append((x, txt))
                current_y = y if current_y is None else current_y
            else:
                current_words.sort(key=lambda t: t[0])
                lines.append(" ".join(wt for _, wt in current_words))
                current_words = [(x, txt)]
                current_y = y
        if current_words:
            current_words.sort(key=lambda t: t[0])
            lines.append(" ".join(wt for _, wt in current_words))

        cell = ws.cell(row=ri + 1, column=ci + 1)
        cell.value = "\n".join(lines)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(XLSX_PATH)
    print(f"created: {XLSX_PATH}")
